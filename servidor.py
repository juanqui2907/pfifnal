"""
TerraShield — Servidor Flask
Soporta API de cálculo IEEE 80-2013, gestión de historiales y exportación.
"""
import os
import sys
import json
import datetime
import webbrowser
import threading

from flask import Flask, request, jsonify, send_from_directory

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

def calcular_malla(p):
    import importlib, calculos_malla
    importlib.reload(calculos_malla)
    return calculos_malla.calcular_malla(p)

def listas_d(*args, **kwargs):
    import importlib, calculos_malla
    importlib.reload(calculos_malla)
    return calculos_malla.listas_d(*args, **kwargs)

PORT = 8000
BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
HISTORIALES_DIR = os.path.join(BASE_DIR, 'historiales')

app = Flask(__name__, static_folder=BASE_DIR, static_url_path='')

# ─── Crear carpeta historiales/ si no existe ──────────────────────────────────
os.makedirs(HISTORIALES_DIR, exist_ok=True)

# ─── Base de datos SQLite de suelos ───────────────────────────────────────────
# suelos.db se genera localmente con convertir_suelos.py y se sube al repo.
# El servidor abre una conexión por hilo (check_same_thread=False) y consulta
# solo los polígonos cuyo bounding box toca el punto. Usa ~5 MB de RAM.
_DB_PATH = os.path.join(BASE_DIR, 'suelos.db')
_db_disponible = os.path.exists(_DB_PATH)

if _db_disponible:
    print(f'[suelos] suelos.db encontrado — endpoint /suelo activo ✓', flush=True)
else:
    print('[suelos] suelos.db no encontrado — endpoint /suelo deshabilitado', flush=True)

def _get_db():
    """Devuelve una conexión SQLite (una por hilo, reutilizable)."""
    import threading as _th
    local = _get_db._local
    if not hasattr(local, 'con'):
        local.con = sqlite3.connect(_DB_PATH, check_same_thread=False)
    return local.con

_get_db._local = __import__('threading').local()

def _ray_cast(point, ring):
    px, py = point
    inside = False
    n = len(ring)
    j = n - 1
    for i in range(n):
        xi, yi = ring[i]
        xj, yj = ring[j]
        if ((yi > py) != (yj > py)) and (px < (xj - xi) * (py - yi) / (yj - yi) + xi):
            inside = not inside
        j = i
    return inside

def _point_in_geojson(lon, lat, geometry):
    px, py = lon, lat

    def in_poly(coords):
        if not _ray_cast((px, py), coords[0]):
            return False
        for hole in coords[1:]:
            if _ray_cast((px, py), hole):
                return False
        return True

    gt = geometry.get('type')
    if gt == 'Polygon':
        return in_poly(geometry['coordinates'])
    if gt == 'MultiPolygon':
        return any(in_poly(c) for c in geometry['coordinates'])
    return False

def _query_suelo(lat, lon):
    """
    Consulta suelos.db filtrando primero por bounding box (índice SQLite),
    luego hace ray-casting solo sobre los candidatos (~5-20 polígonos).
    Usa memoria mínima, no carga nada en RAM al arrancar.
    """
    if not _db_disponible:
        return None

    con = _get_db()
    cur = con.execute(
        '''SELECT tipo_suelo, subtipo_nc, categoria, rho, confianza, geometry
           FROM poligonos
           WHERE min_lon <= ? AND max_lon >= ?
             AND min_lat <= ? AND max_lat >= ?''',
        (lon, lon, lat, lat)
    )
    for row in cur:
        tipo_suelo, subtipo_nc, categoria, rho, confianza, geom_json = row
        geom = json.loads(geom_json)
        if _point_in_geojson(lon, lat, geom):
            return {
                'tipo_suelo': tipo_suelo,
                'subtipo_nc': subtipo_nc,
                'categoria':  categoria,
                'rho':        rho,
                'confianza':  confianza,
            }
    return None

# ─── Helpers de historial ─────────────────────────────────────────────────────

def _safe_nombre(nombre):
    """Nombre de archivo seguro: alfanumérico, guiones y guiones bajos."""
    safe = "".join(c for c in nombre if c.isalnum() or c in (' ', '-', '_')).strip()
    return safe.replace(' ', '_') or 'proyecto'

def _hist_path(nombre):
    return os.path.join(HISTORIALES_DIR, _safe_nombre(nombre) + '.json')

def _load_hist(nombre):
    path = _hist_path(nombre)
    if not os.path.exists(path):
        return []
    with open(path, encoding='utf-8') as f:
        return json.load(f)

def _save_hist(nombre, hist):
    with open(_hist_path(nombre), 'w', encoding='utf-8') as f:
        json.dump(hist, f, ensure_ascii=False, indent=2)

# ─── Archivos estáticos ───────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/<path:filename>')
def static_files(filename):
    return send_from_directory(BASE_DIR, filename)

# ─── API: Gestión de proyectos ────────────────────────────────────────────────

@app.route('/proyectos', methods=['GET'])
def list_proyectos():
    """Lista los nombres de proyectos existentes (archivos .json en historiales/)."""
    files = [f[:-5] for f in os.listdir(HISTORIALES_DIR) if f.endswith('.json')]
    return jsonify(sorted(files))

@app.route('/proyectos/nuevo', methods=['POST'])
def nuevo_proyecto():
    """Crea el archivo de historial vacío para un proyecto si no existe."""
    data   = request.get_json(force=True, silent=True) or {}
    nombre = data.get('nombre', '').strip()
    if not nombre:
        return jsonify({'error': 'nombre requerido'}), 400
    path = _hist_path(nombre)
    if not os.path.exists(path):
        _save_hist(nombre, [])
    return jsonify({'ok': True, 'archivo': os.path.basename(path)})

@app.route('/proyectos/<nombre>', methods=['GET'])
def get_proyecto(nombre):
    """Devuelve el historial completo de un proyecto."""
    return jsonify(_load_hist(nombre))

# ─── API: /espaciamientos ─────────────────────────────────────────────────────

@app.route('/espaciamientos', methods=['GET'])
def get_espaciamientos():
    try:
        forma  = int(request.args.get('forma', 1))
        Lx     = float(request.args.get('Lx',   0))
        Ly     = float(request.args.get('Ly',   0))
        L      = float(request.args.get('L',    0))
        Lb     = float(request.args.get('Lb',   0))
        Li     = float(request.args.get('Li',   0))
        La     = float(request.args.get('La',   0))
        Lb_e   = float(request.args.get('Lb_e', 0))
        Lc_e   = float(request.args.get('Lc_e', 0))
    except (ValueError, TypeError) as e:
        return jsonify({"error": f"Parámetro inválido: {e}"}), 400

    resultado = listas_d(forma, Lx=Lx, Ly=Ly, L=L, Lb=Lb, Li=Li,
                         La=La, Lb_e=Lb_e, Lc_e=Lc_e)
    return jsonify(resultado)

# ─── API: /calcular ───────────────────────────────────────────────────────────

@app.route('/calcular', methods=['POST'])
def post_calcular():
    import traceback as _tb

    # ── DIAGNÓSTICO 1: body completo recibido ─────────────────────────────────
    data = request.get_json(force=True, silent=True)
    print("=" * 60)
    print("[DEBUG /calcular] Body recibido:")
    print(f"  keys       : {list(data.keys()) if data else 'None'}")
    print(f"  'proyecto' : {repr(data.get('proyecto')) if data else 'N/A'}")
    print("=" * 60)

    if not data:
        return jsonify({"error": "Body JSON requerido"}), 400

    required = ["forma", "rho", "h", "Lr", "b", "uv",
                "material_id", "I_falla", "tc", "tf", "XR", "Sf", "Tamb",
                "hs", "ts"]
    missing = [f for f in required if f not in data]
    if missing:
        return jsonify({"error": f"Faltan campos: {', '.join(missing)}"}), 400

    try:
        resultado = calcular_malla(data)
    except Exception as e:
        print("[ERROR /calcular]", _tb.format_exc())
        return jsonify({"error": str(e)}), 500

    # ── DIAGNÓSTICO 2: existencia de carpeta historiales/ ─────────────────────
    print(f"[DEBUG] HISTORIALES_DIR  : {HISTORIALES_DIR}")
    print(f"[DEBUG] carpeta existe   : {os.path.exists(HISTORIALES_DIR)}")
    print(f"[DEBUG] permisos escrit. : {os.access(HISTORIALES_DIR, os.W_OK)}")

    # Crear carpeta si falta (defensa extra)
    os.makedirs(HISTORIALES_DIR, exist_ok=True)

    # ── Guardar iteración en historial ────────────────────────────────────────
    proyecto = data.get('proyecto', '').strip()

    if not proyecto:
        print("[DEBUG] 'proyecto' vacío — historial NO guardado")
    else:
        hist_path = _hist_path(proyecto)
        print(f"[DEBUG] Ruta del JSON   : {hist_path}")
        print(f"[DEBUG] Archivo existe  : {os.path.exists(hist_path)}")

        try:
            # Crear archivo si no existe
            if not os.path.exists(hist_path):
                _save_hist(proyecto, [])
                print(f"[DEBUG] Archivo creado vacío: {hist_path}")

            hist = _load_hist(proyecto)
            print(f"[DEBUG] Iteraciones previas: {len(hist)}")

            sec3 = resultado.get('sec3', {})
            sec5 = resultado.get('sec5', {})
            sec7 = resultado.get('sec7', {})
            sec8 = resultado.get('sec8', {})
            sup  = resultado.get('superficie', {})

            Dx = float(data.get('Dx') or 0)
            Dy = float(data.get('Dy') or 0)
            D_otros = data.get('D') or data.get('Db') or data.get('D1') or data.get('D1_esc')
            if D_otros:
                D = float(D_otros)
            elif Dx > 0 and Dy > 0:
                D = round((Dx + Dy) / 2, 4)
            else:
                D = Dx or Dy

            iteracion = {
                'iter':      len(hist) + 1,
                'timestamp': datetime.datetime.now().isoformat(timespec='seconds'),
                'Lx':        data.get('Lx', data.get('L', 0)),
                'Ly':        data.get('Ly', data.get('L', 0)),
                'D':         round(float(D), 4) if D else 0,
                'h':         data.get('h', 0),
                'Lr':        data.get('Lr', 0),
                'Nr':        sec3.get('nR', 0),
                'hs':        data.get('hs', 0),
                'rho_s':     round(float(sup.get('rho_s', 0)), 2),
                'Rg':        round(float(sec5.get('Rg', 0)), 4),
                'GPR':       round(float(sec7.get('GPR_V', 0)), 2),
                'Em':        round(float(sec8.get('Em_V', 0)), 2),
                'Es':        round(float(sec8.get('Es_V', 0)), 2),
                'ok_Rg':     float(sec5.get('Rg', 999)) <= 1.0,
                'ok_GPR':    bool(sec7.get('gpr_ok', False)),
                'ok_Em':     bool(sec8.get('em_ok', False)),
                'ok_Es':     bool(sec8.get('es_ok', False)),
                'payload':   data,
            }

            # ── DIAGNÓSTICO 3: justo antes y después de escribir ──────────────
            print(f"[DEBUG] Guardando iteración #{iteracion['iter']}...")
            hist.append(iteracion)
            _save_hist(proyecto, hist)
            print(f"[DEBUG] Guardado exitoso — {hist_path}")

            # ── DIAGNÓSTICO 4: verificar que el archivo tiene contenido ───────
            size = os.path.getsize(hist_path)
            print(f"[DEBUG] Tamaño del archivo tras escritura: {size} bytes")

        except Exception as e:
            print(f"[ERROR] Fallo al guardar historial: {e}")
            print(_tb.format_exc())

    return jsonify(resultado)

# ─── API: Exportar Excel ──────────────────────────────────────────────────────

@app.route('/proyectos/<nombre>/exportar/excel', methods=['GET'])
def exportar_excel(nombre):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from flask import send_file
        import io
    except ImportError:
        return jsonify({'error': 'Instale openpyxl: pip install openpyxl'}), 500

    hist = _load_hist(nombre)
    aprobada = next(
        (it for it in reversed(hist)
         if it.get('ok_Rg') and it.get('ok_GPR') and it.get('ok_Em') and it.get('ok_Es')),
        None
    )
    if not aprobada:
        return jsonify({'error': 'No existe ninguna iteración con todos los criterios aprobados'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = 'Diseño Final'

    header_fill = PatternFill('solid', fgColor='0A3D91')
    header_font = Font(color='FFFFFF', bold=True)
    green_fill  = PatternFill('solid', fgColor='C6EFCE')
    center      = Alignment(horizontal='center')

    headers = ['#', 'Fecha', 'Lx (m)', 'Ly (m)', 'D (m)', 'h (m)',
               'Lr (m)', 'Nr', 'hs (m)', 'ρs (Ω·m)',
               'Rg (Ω)', 'GPR (V)', 'Em (V)', 'Es (V)',
               'ok Rg', 'ok GPR', 'ok Em', 'ok Es']
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    fila = [
        aprobada.get('iter'), aprobada.get('timestamp'),
        aprobada.get('Lx'), aprobada.get('Ly'), aprobada.get('D'),
        aprobada.get('h'), aprobada.get('Lr'), aprobada.get('Nr'),
        aprobada.get('hs'), round(aprobada.get('rho_s', 0), 2),
        aprobada.get('Rg'), aprobada.get('GPR'),
        aprobada.get('Em'), aprobada.get('Es'),
        '✔', '✔', '✔', '✔',
    ]
    ws.append(fila)
    for col in range(15, 19):
        ws.cell(row=2, column=col).fill = green_fill

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"TerraShield_{_safe_nombre(nombre)}.xlsx"
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ─── API: Exportar PDF ────────────────────────────────────────────────────────

@app.route('/proyectos/<nombre>/exportar/pdf', methods=['GET', 'POST'])
def exportar_pdf(nombre):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib import colors
        from flask import send_file
        import io, base64
    except ImportError:
        return jsonify({'error': 'Instale reportlab: pip install reportlab'}), 500

    # Recibir imagen de la vista 2D si viene en el body (POST)
    malla2d_img = None
    if request.method == 'POST':
        body = request.get_json(silent=True) or {}
        malla2d_img = body.get('malla2d_img')  # data URL base64 PNG

    hist = _load_hist(nombre)
    aprobada = next(
        (it for it in reversed(hist)
         if it.get('ok_Rg') and it.get('ok_GPR') and it.get('ok_Em') and it.get('ok_Es')),
        None
    )
    if not aprobada:
        return jsonify({'error': 'No existe ninguna iteración con todos los criterios aprobados'}), 400

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=40, bottomMargin=40,
                            leftMargin=50, rightMargin=50)
    styles = getSampleStyleSheet()
    story  = []

    story.append(Paragraph('TerraShield — Malla de puesta a tierra IEEE 80-2013', styles['Title']))
    story.append(Paragraph(f'Proyecto: <b>{nombre}</b>', styles['Normal']))
    story.append(Paragraph(
        f'Fecha de exportación: {datetime.datetime.now().strftime("%d/%m/2024 %H:%M")}',
        styles['Normal']
    ))
    story.append(Spacer(1, 20))

    # Insertar visualización 2D si fue enviada desde el cliente
    if malla2d_img:
        try:
            img_data  = malla2d_img.split(',', 1)[-1]
            img_bytes = base64.b64decode(img_data)
            img_buf   = io.BytesIO(img_bytes)
            img       = Image(img_buf, width=450, height=250)
            img.hAlign = 'CENTER'
            story.append(Paragraph('Visualización de la malla — Vista 2D', styles['Heading2']))
            story.append(Spacer(1, 8))
            story.append(img)
            story.append(Spacer(1, 20))
        except Exception:
            pass  # Si falla la imagen, continúa sin ella

    story.append(Paragraph('Diseño Final Aprobado', styles['Heading2']))
    story.append(Spacer(1, 8))

    nd = lambda k: str(aprobada.get(k, 'N/D'))  # helper: valor o N/D

    tabla_data = [
        ['Parametro',                  'Valor',                                        'Unidad'],
        ['Iteracion N.',               nd('iter'),                                     ''],
        ['Fecha de calculo',           aprobada.get('timestamp', 'N/D'),               ''],
        ['Lx - longitud X',            nd('Lx'),                                       'm'],
        ['Ly - longitud Y',            nd('Ly'),                                       'm'],
        ['D - espaciamiento',          nd('D'),                                        'm'],
        ['h - profundidad de malla',   nd('h'),                                        'm'],
        ['Lr - longitud de varilla',   nd('Lr'),                                       'm'],
        ['Nr - numero de varillas',    nd('Nr'),                                       ''],
        ['hs - capa superficial',      nd('hs'),                                       'm'],
        ['rhos - resistividad sup.',   str(round(aprobada.get('rho_s', 0), 2)),        'Ohm.m'],
        ['Rg - resist. de malla',      nd('Rg'),                                       'Ohm [OK]'],
        ['GPR - potencial de tierra',  nd('GPR'),                                      'V   [OK]'],
        ['Em - voltaje de malla',      nd('Em'),                                       'V   [OK]'],
        ['Es - voltaje de paso',       nd('Es'),                                       'V   [OK]'],
    ]

    t = Table(tabla_data, colWidths=[210, 110, 80])
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0),  (-1, 0),  colors.HexColor('#0A3D91')),
        ('TEXTCOLOR',     (0, 0),  (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0),  (-1, -1), 'Helvetica'),
        ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
        ('ROWBACKGROUNDS',(0, 1),  (-1, -1), [colors.white, colors.HexColor('#EEF3FF')]),
        ('GRID',          (0, 0),  (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('FONTSIZE',      (0, 0),  (-1, -1), 9),
        ('TOPPADDING',    (0, 0),  (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0),  (-1, -1), 5),
        ('LEFTPADDING',   (0, 0),  (-1, -1), 8),
    ]))
    story.append(t)

    try:
        doc.build(story)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error al construir el PDF: {str(e)}'}), 500

    output.seek(0)

    fname = f"TerraShield_{_safe_nombre(nombre)}.pdf"
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype='application/pdf')

# ─── API: Tipo de suelo IEEE 80 por coordenadas ───────────────────────────────

@app.route('/suelo', methods=['GET'])
def get_suelo():
    """
    Consulta el tipo de suelo IEEE 80 para un punto lat/lon.
    El JSON (194 MB) se carga una sola vez en memoria al arrancar el servidor.
    El cliente recibe solo el resultado del polígono que contiene el punto.

    Query params:
        lat  — latitud  decimal (ej. 10.39)
        lon  — longitud decimal (ej. -75.47)

    Respuesta:
        200 + {tipo_suelo, subtipo_nc, categoria, rho, confianza}
        204 (sin contenido) si el punto no cae en ningún polígono
        503 si el índice aún no terminó de cargar
    """
    try:
        lat = float(request.args['lat'])
        lon = float(request.args['lon'])
    except (KeyError, ValueError):
        return jsonify({'error': 'Parámetros lat y lon requeridos (decimales)'}), 400

    if not _db_disponible:
        return jsonify({'error': 'Base de datos de suelos no disponible'}), 503

    resultado = _query_suelo(lat, lon)
    if resultado is None:
        return ('', 204)   # punto fuera de cobertura

    return jsonify(resultado)

# ─── API: Apantallamiento ─────────────────────────────────────────────────────

@app.route('/apantallamiento/calcular', methods=['POST'])
def post_calcular_apantallamiento():
    import traceback as _tb
    data = request.get_json(force=True, silent=True)
    if not data:
        return jsonify({'error': 'Body JSON requerido'}), 400
    if not data.get('masts'):
        return jsonify({'error': 'Se requiere al menos un mástil'}), 400
    try:
        import importlib, calculos_apant
        importlib.reload(calculos_apant)
        resultado = calculos_apant.calcular_apantallamiento(data)
        return jsonify({'ok': True, **resultado})
    except Exception as e:
        print('[ERROR /apantallamiento/calcular]', _tb.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/apantallamiento/recomendar', methods=['POST'])
def post_recomendar_apantallamiento():
    import traceback as _tb
    data = request.get_json(force=True, silent=True)
    if not data:
        return jsonify({'error': 'Body JSON requerido'}), 400
    try:
        import importlib, calculos_apant
        importlib.reload(calculos_apant)
        params       = data.get('params', {})
        S            = float(data.get('S') or 0)
        verification = data.get('verification', [])
        if not params.get('masts'):
            return jsonify({'error': 'Se requiere al menos un mástil en params'}), 400
        if S <= 0:
            return jsonify({'error': 'Se requiere S > 0'}), 400
        recs = calculos_apant.calcular_recomendaciones(params, S, verification)
        return jsonify({'ok': True, 'recommendations': recs})
    except Exception as e:
        print('[ERROR /apantallamiento/recomendar]', _tb.format_exc())
        return jsonify({'error': str(e)}), 500

# ─── Arranque ─────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print("=" * 46)
    print("  TERRASHIELD v1.0.0")
    print(f"  Servidor Flask en http://localhost:{PORT}")
    print("  Cierre esta ventana para detener.")
    print("=" * 46)
    app.run(host='0.0.0.0', port=PORT, debug=False)
